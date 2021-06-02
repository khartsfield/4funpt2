"""
Reference: https://medium.com/code-85/how-to-request-command-line-input-in-python-80f45e9032fe#:~:
text=Fortunately%20for%20new%20Python%20enthusiasts,and%20submit%20by%20pressing%20ENTER
"""

from datetime import date, datetime
from openpyxl import Workbook  # pip3
from openpyxl import load_workbook
import os
import re


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False


def excel_create(file_name, log_list, time_list1, temp_list, time_list2, r1, g1, b1, time_list3, r2, g2, b2, time_list4, conductivity_list, pump_list):
    # change the input_list to specific data name later
    workbook = Workbook()
    sheet = workbook.active
    # data labels write in column A of the Excel file
    sheet["A1"] = "Lab Name"
    sheet["A2"] = "Time"
    sheet["A3"] = "Beaker's Capacity (mL)"
    sheet["A4"] = "Density of Solvent (g/mL)"
    sheet["A5"] = "Volume of Liquid Solvent (mL)"
    sheet["A6"] = "Mass of Solvent (g)"
    sheet["A7"] = "Mass of final solution"
    sheet["A8"] = "Concentration (% by mass)"

    sheet["C1"] = "Time of temperature sensor (s)"
    sheet["D1"] = "Temperature data (F)"
    sheet["E1"] = "Time of RGB sensor 1 (s)"
    sheet["F1"] = "RGB sensor 1 Red data"
    sheet["G1"] = "RGB sensor 1 Green data"
    sheet["H1"] = "RGB sensor 1 Blue data"
    sheet["I1"] = "Time of RGB sensor 2 (s)"
    sheet["J1"] = "RGB sensor 2 Red data"
    sheet["K1"] = "RGB sensor 2 Green data"
    sheet["L1"] = "RGB sensor 2 Blue data"
    sheet["M1"] = "Time of conductivity sensor (s)"
    sheet["N1"] = "Conductivity data (μS/cm)"
    sheet["O1"] = "solvent data from the pump"

    # data list input by the user
    n = 1
    for list_item in log_list:
        cell_name = "B{}".format(n)  # for horizontal column B
        sheet[cell_name] = list_item  # data input starts from B1
        n += 1

    cell_number = 2
    for input in time_list1:  # for loop
        cell_name = "C{}".format(cell_number)  # for horizontal column C
        sheet[cell_name] = input  # data input starts from C2
        cell_number += 1

    cell_number = 2
    for input in temp_list:  # for loop
        cell_name = "D{}".format(cell_number)  # for horizontal column D, data input starts from D2
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in time_list2:
        cell_name = "E{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in r1:
        cell_name = "F{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in g1:
        cell_name = "G{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in b1:
        cell_name = "H{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in time_list3:
        cell_name = "I{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in r2:
        cell_name = "J{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in g2:
        cell_name = "K{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in b2:
        cell_name = "L{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in time_list4:
        cell_name = "M{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    cell_number = 2
    for input in conductivity_list:
        cell_name = "N{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1


    cell_number = 2
    for input in pump_list:
        cell_name = "O{}".format(cell_number)
        sheet[cell_name] = input
        cell_number += 1

    workbook.save(filename=file_name)  # save file


def info_input():
    # info input function, lab name needs to contains alphabets
    while True:
        lab_name = input("Enter your lab's name: ")
        print(lab_name)
        # re.findall('[^a-zA-Z]') limits the input to be only alphabets
        if not re.findall('[^a-zA-Z]', lab_name):
            print("Error: name must be alphabetic characters: ")
        else:
            break

# show the current date and time
    current_time = datetime.now()
    print(current_time)

# use if, else if, else to limit the input values to be positive and smaller than a certain number
    while True:
        print("enter integer only")
        capacity = input("Enter beaker's capacity (unit in mL): ")
        if is_number(capacity):
            capacity = int(capacity)
            print(capacity, "mL")
            if capacity < 0:
                print("Error: beaker's capacity cannot be negative")
            elif capacity > 1000:
                print("Error: beaker's capacity is too large")
            else:
                break
        else:
            print("Error: Input data needs to be number")

    while True:
        density = input("Enter the density of solvent (unit in g/mL): ")
        if is_number(density):
            density = float(density)
            print(density, "g/mL")
            if density < 0:
                print("Error: density value cannot be negative")
            elif density > 100:
                print("Error: density value is too large")
            else:
                break
        else:
            print("Error: Input data needs to be number")

    data_log = [lab_name, current_time, capacity, density]
    return data_log


def data_check1(list_item1):
    while True:
        print(list_item1)
        print("Are all the input information correct? ")
        ans = str(input("Enter Y, y, Yes, YES, or yes for yes; N, n, No, NO, or no for no: "))
        if ans == "Y" or ans == "Yes" or ans == "yes" or ans == "YES" or ans == "y":
            break
        elif ans == "N" or ans == "No" or ans == "NO" or ans == "no" or ans == "n":
            while True:
                print("0 is for lab name, 1 is for current time, 2 is for beaker's capacity, etc ")
                position = input("position number in the list: ")
                if is_number(position):
                    position = int(position)
                    if position < 0:
                        print("Error: position number cannot be negative ")
                    elif position > 3:
                        print("Error: position number is too large, does not exist in the data list ")
                    else:
                        break
                else:
                    print("Error: Input data needs to be number")
            if list_item1[position] == list_item1[0]:
                while True:
                    value = input("Enter correct lab's name: ")
                    print(value)
                    if not re.findall('[^a-zA-Z]', value):
                        print("Error: name must be alphabetic characters: ")
                    else:
                        list_item1[position] = value
                        break
            else:
                while True:
                    value = input("Enter correct information: ")
                    if is_number(value):
                        value = float(value)
                        if value < 0:
                            print("Error: value cannot be negative ")
                        elif value > 1000:
                            print("Error: value is too large")
                        else:
                            print(value)
                            list_item1[position] = value
                            break
                    else:
                        print("Error: Input data needs to be number")
        else:
            print("Please only enter Y, y, Yes, YES, or yes for yes; N, n, No, NO, or no for no")


def data_check2(list_item2):
    while True:
        print(list_item2)
        print("Are all the input information correct? ")
        ans = str(input("Enter Y, y, Yes, YES, or yes for yes; N, n, No, NO, or no for no: "))
        if ans == "Y" or ans == "Yes" or ans == "yes" or ans == "YES" or ans == "y":
            break
        elif ans == "N" or ans == "No" or ans == "NO" or ans == "no" or ans == "n":
            while True:
                print("0 is for mass of solute, 1 is for solution mass, 2 is for concentration, etc ")
                position = input("position number in the list: ")
                if is_number(position):
                    position = int(position)
                    if position < 0:
                        print("Error: position number cannot be negative ")
                    elif position > 2:
                        print("Error: position number is too large, does not exist in the data list ")
                    else:
                        break
                else:
                    print("Error: Input data needs to be number")
            while True:
                value = input("Enter correct information: ")
                if is_number(value):
                    value = float(value)
                    if value < 0:
                        print("Error: value cannot be negative ")
                    elif value > 1000:
                        print("Error: value is too large")
                    else:
                        print(value)
                        list_item2[position] = value
                        break
                else:
                    print("Error: Input data needs to be number")
        else:
            print("Please only enter Y, y, Yes, YES, or yes for yes; N, n, No, NO, or no for no")


def excel_add(filename, v_solute, m_solvent, mass_solution, concentration):
    # to add the data list after both RGB sensors detect the saturation point reached
    info_add = [v_solute, m_solvent, mass_solution, concentration]
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    cell = 5
    for list_item in info_add:
        cell_name = "B{}".format(cell)
        sheet[cell_name] = list_item
        cell += 1
    workbook.save(filename=filename)


def excel_placement():
    print("Input the Excel name you want your Excel file to have, remember to include .xlsx or .csv")
    file_name = input("Enter your Excel file name: ")
    path = file_name
    print("Your local path is" + os.getcwd())
    while True:
        print("If you have already answered the question of saving in other location once, ")
        print("enter N to save the file in the location you chose")
        print("Enter Y, y, Yes, YES, or yes for yes; N, n, No, NO, or no for no ")
        selection = input("Do you want to save in other location? : ")
        if selection == "Y" or selection == "Yes" or selection == "yes" or selection == "YES" or selection == "y":
            print("Please go to the folder you want the Excel to be located, and copy the address from the address bar.")
            path = input("Enter your path: ")
            if os.path.isdir(path):
                path = path+"/"+file_name
                print("Your selected location: ", path)
            else:
                print("Error: enter the correct path")
        elif selection == "N" or selection == "No" or selection == "NO" or selection == "no" or selection == "n":
            print("file will save in ", path)
            return path
            # break
        else:
            print("Enter only Enter Y, y, Yes, YES, or yes for yes; N, n, No, NO, or no for no ")
